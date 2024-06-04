import openpyxl
import random
from openpyxl import Workbook, load_workbook

root_path = '/content/drive/MyDrive/final_project/model_test/kobert/data'

# 카테고리(이름) + 문장
def wellness_question_data():
    wellness_file = root_path + "/배경데이터 1차 통합.xlsx"
    wellness_q_output = root_path + "/wellness_dialog_question_배경.txt"

    f = open(wellness_q_output, 'w')
    wb = load_workbook(filename=wellness_file)
    ws = wb[wb.sheetnames[0]]

    first_line = 0
    for row in ws.iter_rows():
        if first_line == 0:
            first_line = 1
            continue
        f.write(row[0].value + "    " + row[1].value + "\n")

    f.close()

# 카테고리(이름 + 숫자)
def wellness_category_data():
    wellness_file = root_path + "/배경데이터 1차 통합.xlsx"
    wellness_c_output = root_path + "/wellness_dialog_category_배경.txt"

    f = open(wellness_c_output, 'w')
    wb = load_workbook(filename=wellness_file)
    ws = wb[wb.sheetnames[0]]

    first_line = 0
    category_count = 0
    cate_dict = []
    for row in ws.iter_rows():
        if first_line == 0:
            first_line = 1
            continue
        a = row[0].value
        if a not in cate_dict:
            cate_dict.append(a)
            f.write(row[0].value + "    " + str(category_count) + "\n")
            category_count += 1

    f.close()

# 문장 + 카테고리(숫자)
def wellness_text_classification_data():
    wellness_category_file = root_path + "/wellness_dialog_category_배경.txt"
    wellness_question_file = root_path + "/wellness_dialog_question_배경.txt"
    wellness_text_classification_file = root_path + "/wellness_dialog_for_text_classification_배경.txt"

    cate_file = open(wellness_category_file, 'r', encoding='utf-8')
    ques_file = open(wellness_question_file, 'r', encoding='utf-8')
    text_classfi_file = open(wellness_text_classification_file, 'w', encoding='utf-8')

    category_lines = cate_file.readlines()
    cate_dict = {}
    for line_num, line_data in enumerate(category_lines):
        data = line_data.split('    ')
        cate_dict[data[0]] = data[1][:-1]

    ques_lines = ques_file.readlines()
    ques_dict = {}
    for line_num, line_data in enumerate(ques_lines):
        data = line_data.split('    ')
        text_classfi_file.write(data[1][:-1] + "    " + cate_dict[data[0]] + "\n")

    cate_file.close()
    ques_file.close()
    text_classfi_file.close()

# train, test 데이터 분리
def seperate_wellness_data():
    file_path = root_path + "/wellness_dialog_for_text_classification_배경.txt"
    train_file_path = root_path + "/wellness_dialog_for_text_classification_train_배경.txt"
    test_file_path = root_path + "/wellness_dialog_for_text_classification_test_배경.txt"

    sperated_file = open(file_path, 'r')
    train_file = open(train_file_path, 'w')
    test_file = open(test_file_path, 'w')

    sperated_file_lines = sperated_file.readlines()
    ques_dict = {}
    for line_num, line_data in enumerate(sperated_file_lines):
        rand_num = random.randint(0, 10)
        if rand_num < 10:
            train_file.write(line_data)
        else:
            test_file.write(line_data)

    sperated_file.close()
    train_file.close()
    test_file.close()

if __name__ == '__main__':
    wellness_question_data()
    # wellness_answer_data()
    wellness_category_data()
    wellness_text_classification_data()
    seperate_wellness_data()