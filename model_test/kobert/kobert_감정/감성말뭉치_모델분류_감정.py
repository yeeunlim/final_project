import os
import torch
from transformers import AdamW, BertModel
from model.classifier import BERTClassifier, kobert_input
from model.dataloader import WellnessTextClassificationDataset
from kobert_tokenizer import KoBERTTokenizer
import torch.nn as nn
import random
import openpyxl
from openpyxl import Workbook, load_workbook

# 감정 카테고리 로드
def load_emotion_category():
    # 카테고리 파일 경로 설정
    category_path = './data/wellness_dialog_category_감정.txt'

    # 카테고리 파일 열기
    c_f = open(category_path, 'r')
    # 파일의 모든 라인 읽기
    category_lines = c_f.readlines()

    # 각 라인을 처리하여 딕셔너리에 저장
    category = {}
    for line_num, line_data in enumerate(category_lines):
        data = line_data.split('    ')
        category[data[1][:-1]] = data[0]
    return category

if __name__ == "__main__":
    # 디바이스 설정
    ctx = "cuda" if torch.cuda.is_available() else "cpu"
    device = torch.device(ctx)
    # 체크포인트 로드
    checkpoint_path = "./checkpoint"
    save_ckpt_path = checkpoint_path + "/kobert-wellness-text-classification_감정.pth"
    checkpoint = torch.load(save_ckpt_path, map_location=device)
    # KoBERT 모델 로드
    bertmodel = BertModel.from_pretrained('skt/kobert-base-v1', return_dict=False)
    # 모델 초기화
    model = BERTClassifier(bertmodel)
    # 모델 상태 딕셔너리 로드
    model.load_state_dict(checkpoint['model_state_dict'])
    # 모델을 디바이스에 로드
    model.to(ctx)
    model.eval()

    # 토크나이저 로드
    tokenizer = KoBERTTokenizer.from_pretrained('skt/kobert-base-v1')

    # 감정 카테고리 로드
    category = load_emotion_category()

    # 감성 대화 말뭉치 로드
    corpus_file = './data/감성대화말뭉치_감정분류.xlsx'
    wb = load_workbook(filename=corpus_file)
    ws = wb[wb.sheetnames[0]]
    first_line = 0
    # 각 행을 순회
    for row in ws.iter_rows():
        if first_line == 0:
            first_line = 1
            continue
        # 문장 데이터 가져오기
        sent = row[2].value
        # 입력 데이터 생성
        data = kobert_input(tokenizer, sent, device, 512)
        # 모델에 입력하여 출력값 얻기
        output = model(**data)

        # 로짓 값 추출
        logit = output[0]
        # 소프트맥스 적용
        softmax_logit = torch.softmax(logit, dim=-1)
        softmax_logit = softmax_logit.squeeze()

        # 최댓값 인덱스 및 값 추출
        max_index = torch.argmax(softmax_logit).item()
        max_index_value = softmax_logit[torch.argmax(softmax_logit)].item()

        # 예측된 카테고리 추출
        cate_pred = category[str(max_index)]

        # 예측된 카테고리를 4번째 열에 저장
        row[3].value = cate_pred

    # 변경된 내용을 파일에 저장
    wb.save(corpus_file)